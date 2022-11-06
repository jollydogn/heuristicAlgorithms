############################################################################

# Created by: Prof. Valdecy Pereira, D.Sc.
# UFF - Universidade Federal Fluminense (Brazil)
# email:  valdecy.pereira@gmail.com
# Metaheuristic: Simulated Annealing

# PEREIRA, V. (2022). GitHub repository: https://github.com/Valdecy/pyMetaheuristic

############################################################################

# Required Libraries
import math
import numpy as np
import random
import os
from openpyxl import Workbook, load_workbook
import array

############################################################################

# Function
def target_function():
    return

############################################################################

# Function: Initialize Variables
def initial_guess(min_values = [-5,-5], max_values = [5,5], target_function = target_function):
    n = 1
    guess = np.zeros((n, len(min_values) + 1))
    for j in range(0, len(min_values)):
         guess[0,j] = random.uniform(min_values[j], max_values[j]) 
    guess[0,-1] = target_function(guess[0,0:guess.shape[1]-1])
    return guess

# Function: Epson Vector
def epson_vector(guess, mu = 0, sigma = 1):
    epson = np.zeros((1, guess.shape[1]-1))
    for j in range(0, guess.shape[1]-1):
        epson[0,j] = float(np.random.normal(mu, sigma, 1))
    return epson

# Function: Updtade Solution
def update_solution(guess, epson, min_values = [-5,-5], max_values = [5,5], target_function = target_function):
    updated_solution = np.copy(guess)
    for j in range(0, guess.shape[1] - 1):
        if (guess[0,j] + epson[0,j] > max_values[j]):
            updated_solution[0,j] = random.uniform(min_values[j], max_values[j])
        elif (guess[0,j] + epson[0,j] < min_values[j]):
            updated_solution[0,j] = random.uniform(min_values[j], max_values[j])
        else:
            updated_solution[0,j] = guess[0,j] + epson[0,j] 
    updated_solution[0,-1] = target_function(updated_solution[0,0:updated_solution.shape[1]-1])
    return updated_solution

############################################################################

# SA Function
# def simulated_annealing(min_values = [-5,-5], max_values = [5,5], mu = 0, sigma = 1, initial_temperature = 1.0, temperature_iterations = 1000, final_temperature = 0.0001, alpha = 0.9, target_function = target_function, verbose = True):    
#     guess = initial_guess(min_values = min_values, max_values = max_values, target_function = target_function)
#     epson = epson_vector(guess, mu = mu, sigma = sigma)
#     best  = np.copy(guess)
#     fx_best = guess[0,-1]
#     Temperature = float(initial_temperature)
#     while (Temperature > final_temperature): 
#         for repeat in range(0, temperature_iterations):
#             if (verbose == True):
#                 print('Temperature = ', round(Temperature, 4), ' ; iteration = ', repeat, ' ; f(x) = ', round(best[0, -1], 4))
#             fx_old    =  guess[0,-1]    
#             epson     = epson_vector(guess, mu = mu, sigma = sigma)
#             new_guess = update_solution(guess, epson, min_values = min_values, max_values = max_values, target_function = target_function)
#             fx_new    = new_guess[0,-1] 
#             delta     = (fx_new - fx_old)
#             r         = int.from_bytes(os.urandom(8), byteorder = "big") / ((1 << 64) - 1)
#             p         = np.exp(-delta/Temperature)
#             print(best)
#             if (delta < 0 or r <= p):
#                 guess = np.copy(new_guess)   
#             if (fx_new < fx_best):
#                 fx_best = fx_new
#                 best    = np.copy(guess)
#         Temperature = alpha*Temperature   
#         return best

# SA Geometrik
def simulated_annealing_geometric(min_values = [-5,-5], max_values = [5,5], mu = 0, sigma = 1, initial_temperature = None, temperature_iterations = 1000, final_temperature = 0.0001, alpha = 0.9, target_function = target_function, verbose = True):    
    guess = initial_guess(min_values = min_values, max_values = max_values, target_function = target_function)
    epson = epson_vector(guess, mu = mu, sigma = sigma)
    best  = np.copy(guess)
    fx_best = guess[0,-1]
    Temperature = float(initial_temperature)
    step = 1
    
    excelstep = []
    exceltemp = []
    excelfx = []
    excelbest = []
    while (Temperature > final_temperature): 
        # for repeat in range(0, temperature_iterations):
        if (verbose == True):
            print('Step = ',step ,' Temperature = ', round(Temperature, 4), ' f(x) = ', round(best[0, -1], 4))
            
            excelstep.append(step)
            exceltemp.append(round(Temperature, 4)) 
            excelfx.append(round(best[0, -1], 4)) 
            
            fx_old    =  guess[0,-1]    
            epson     = epson_vector(guess, mu = mu, sigma = sigma)
            new_guess = update_solution(guess, epson, min_values = min_values, max_values = max_values, target_function = target_function)
            fx_new    = new_guess[0,-1] 
            delta     = (fx_new - fx_old)
            r         = int.from_bytes(os.urandom(8), byteorder = "big") / ((1 << 64) - 1)
            p         = np.exp(-delta/Temperature)
            print('Best solution vector;\n',best,'\n')
            
            excelbest.append(best) 
            step += 1
        if (delta < 0 or r <= p):
                guess = np.copy(new_guess)   
        if (fx_new < fx_best):
                fx_best = fx_new
        best    = np.copy(guess)
        Temperature = alpha*Temperature   
        return best
        
        
    # for i in excelbest:
    #     print(i)
    
        if (initial_temperature == 1000):
            
            wb = load_workbook(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            sheet = wb.active 
            
            sheet.cell(row=1, column=1).value = 'Geometric - Init temp 1000'
            sheet.cell(row=step+1, column=1).value = 'Step'
            sheet.cell(row=step+1, column=2).value = 'Temperature' 
            sheet.cell(row=step+1, column=3).value = 'f(x)' 
            sheet.cell(row=2, column=4).value = 'Best'
            
            for a in range(len(excelstep)):
                temp = excelstep[a]
                sheet.cell(row=a+3, column=1).value = temp
                
            for b in range(len(exceltemp)):
                temp = exceltemp[b]
                sheet.cell(row=b+3, column=2).value = temp
                
            for c in range(len(excelfx)):
                temp = excelfx[c]
                sheet.cell(row=c+3, column=3).value = temp
            
            wb.save(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            
            wb.close()
            
        elif (initial_temperature == 5000):
            
            wb = load_workbook(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            sheet = wb.active 
            
            sheet.cell(row=1, column=7).value = 'Geometric - Init temp 5000'
            sheet.cell(row=step+1, column=7).value = 'Step'
            sheet.cell(row=step+1, column=8).value = 'Temperature' 
            sheet.cell(row=step+1, column=9).value = 'f(x)' 
            sheet.cell(row=2, column=10).value = 'Best'
            
            for a in range(len(excelstep)):
                temp = excelstep[a]
                sheet.cell(row=a+3, column=7).value = temp
                
            for b in range(len(exceltemp)):
                temp = exceltemp[b]
                sheet.cell(row=b+3, column=8).value = temp
                
            for c in range(len(excelfx)):
                temp = excelfx[c]
                sheet.cell(row=c+3, column=9).value = temp
            
            wb.save(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            
            wb.close()
            
        elif (initial_temperature == 10000):
            
            wb = load_workbook(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            sheet = wb.active 
            
            sheet.cell(row=1, column=13).value = 'Geometric - Init temp 10000'
            sheet.cell(row=step+1, column=13).value = 'Step'
            sheet.cell(row=step+1, column=14).value = 'Temperature' 
            sheet.cell(row=step+1, column=15).value = 'f(x)' 
            sheet.cell(row=2, column=16).value = 'Best'
            
            for a in range(len(excelstep)):
                temp = excelstep[a]
                sheet.cell(row=a+3, column=13).value = temp
                
            for b in range(len(exceltemp)):
                temp = exceltemp[b]
                sheet.cell(row=b+3, column=14).value = temp
                
            for c in range(len(excelfx)):
                temp = excelfx[c]
                sheet.cell(row=c+3, column=15).value = temp
            
            wb.save(r'C:/Users/Emirhan/Desktop/OKUL/4.SINIF 1.DÖNEM/INTRODUCTION TO HEURISTIC ALGORITHMS/HAFTA 3/SA - Ödev/SA/Results.xlsx')  
            
            wb.close()

# SA Lineer
def simulated_annealing_linear(min_values = [-5,-5], max_values = [5,5], mu = 0, sigma = 1, initial_temperature = 1000.0, temperature_iterations = 1000, final_temperature = 0.0001, alpha = 0.9, target_function = target_function, verbose = True):    
    guess = initial_guess(min_values = min_values, max_values = max_values, target_function = target_function)
    epson = epson_vector(guess, mu = mu, sigma = sigma)
    best  = np.copy(guess)
    fx_best = guess[0,-1]
    Temperature = float(initial_temperature)
    while (Temperature > final_temperature): 
        # for repeat in range(0, temperature_iterations):
        if (verbose == True):
            print('Temperature = ', round(Temperature, 4), ' f(x) = ', round(best[0, -1], 4))
            fx_old    =  guess[0,-1]    
            epson     = epson_vector(guess, mu = mu, sigma = sigma)
            new_guess = update_solution(guess, epson, min_values = min_values, max_values = max_values, target_function = target_function)
            fx_new    = new_guess[0,-1] 
            delta     = (fx_new - fx_old)
            r         = int.from_bytes(os.urandom(8), byteorder = "big") / ((1 << 64) - 1)
            p         = np.exp(-delta/Temperature)
            print(best)
        if (delta < 0 or r <= p):
                guess = np.copy(new_guess)   
        if (fx_new < fx_best):
                fx_best = fx_new
        best    = np.copy(guess)
        # Temperature = alpha*Temperature   
        # return best
        Temperature = Temperature - alpha
    
############################################################################